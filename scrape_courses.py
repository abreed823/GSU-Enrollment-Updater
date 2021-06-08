import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.pagebreak import RowBreak, ColBreak
from datetime import date
from itertools import chain


class Courses:
    today = None
    book = None
    new_sheet = None
    cell_border = None

    # Declares starting rows for each campus
    alp_row = 5
    clk_row = 42
    dec_row = 80
    dun_row = 118
    newt_row = 156
    onl_row = 195

    # Defines column of each piece of data in data frame
    col_crn = 2
    col_subj = 3
    col_class = 4
    col_sec = 5
    col_campus = 6
    col_credits = 7
    col_title = 8
    col_days = 9
    col_time = 10
    col_cap = 11
    col_act = 12
    col_comments = 17
    col_prof = 18
    col_location = 20

    # Sets repetitive border styles
    regular_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    legend_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    bottom_legend_border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

    def create_new_sheet(self, file):
        # Creates new sheet
        self.today = date.today().strftime('%m-%d-%Y')
        self.book = load_workbook(file)
        self.book.create_sheet('NEW TEST SHEET ' + self.today)
        self.book.save(file)
        self.new_sheet = self.book.worksheets[-1]

        # Sets sheet column widths
        self.new_sheet.column_dimensions['A'].width = 5.83
        self.new_sheet.column_dimensions['B'].width = 6.83
        self.new_sheet.column_dimensions['C'].width = 6
        self.new_sheet.column_dimensions['D'].width = 10.5
        self.new_sheet.column_dimensions['E'].width = 9.5
        self.new_sheet.column_dimensions['F'].width = 30
        self.new_sheet.column_dimensions['G'].width = 5.5
        self.new_sheet.column_dimensions['H'].width = 18.5
        self.new_sheet.column_dimensions['I'].width = 5.5
        self.new_sheet.column_dimensions['J'].width = 6.5
        self.new_sheet.column_dimensions['K'].width = 9.83
        self.new_sheet.column_dimensions['L'].width = 17

        # list(chain(range(1, 5), range(40, 44), range(78, 82), range(116, 120), range(156, 160), range(195, 199)))
        # Appends proper headings to each campus
        for row in [1, 38, 76, 114, 152, 191]:
            self.new_sheet.cell(row=row, column=6, value='Perimeter College').font = Font(bold=True)
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=6, end_column=8)
            row += 1

            self.new_sheet.cell(row=row, column=6, value='HONORS COLLEGE COURSES').font = Font(bold=True)
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=6, end_column=8)
            row += 1

            year = date.today().year
            if date.today().month >= 2 or date.today().month <= 8:
                semester = 'FALL'
            else:
                semester = 'SPRING'

            self.new_sheet.cell(row=row, column=6, value=f'{semester} Semester {year}').font = Font(bold=True)
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=6, end_column=8)

            self.new_sheet.cell(row=row, column=11, value=f"Updated {date.today().strftime('%m/%d/%Y')}").font = Font(size=9)
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=11, end_column=12)

        # Table headers
        for row in [5, 42, 80, 118, 156, 195]:

            self.new_sheet.cell(row=row, column=2, value='CRN').font = Font(size=9, bold=True)
            self.new_sheet.cell(row=row, column=3, value='COURSE ID').font = Font(size=9, bold=True)
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=3, end_column=4)
            self.new_sheet.cell(row=row, column=5, value='CREDITS').font = Font(size=9, bold=True)
            self.new_sheet.cell(row=row, column=6, value='COURSE NAME').font = Font(size=9, bold=True)
            self.new_sheet.cell(row=row, column=7, value='DAY').font = Font(size=9, bold=True)

            # Makes the headers appropriately different for the Online Campus
            if row != 195:
                self.new_sheet.cell(row=row, column=8, value='TIME').font = Font(size=9, bold=True)
                self.new_sheet.cell(row=row, column=9, value='ACT').font = Font(size=9, bold=True)
                self.new_sheet.cell(row=row, column=10, value='CAP').font = Font(size=9, bold=True)
                self.new_sheet.cell(row=row, column=11, value='ROOM').font = Font(size=9, bold=True)
                self.new_sheet.cell(row=row, column=12, value='FACULTY').font = Font(size=9, bold=True)
            else:
                self.new_sheet.cell(row=row, column=8, value='FACULTY').font = Font(size=9, bold=True)
                self.new_sheet.cell(row=row, column=9, value='ACT').font = Font(size=9, bold=True)
                self.new_sheet.cell(row=row, column=10, value='CAP').font = Font(size=9, bold=True)

        # Adds campus names to sheet
        self.new_sheet.cell(row=4, column=2, value='ALPHARETTA CAMPUS').font = Font(size=9, bold=True)
        self.new_sheet.cell(row=41, column=2, value='CLARKSTON CAMPUS').font = Font(size=9, bold=True)
        self.new_sheet.cell(row=79, column=2, value='DECATUR CAMPUS').font = Font(size=9, bold=True)
        self.new_sheet.cell(row=117, column=2, value='DUNWOODY CAMPUS').font = Font(size=9, bold=True)
        self.new_sheet.cell(row=155, column=2, value='NEWTON CAMPUS').font = Font(size=9, bold=True)
        self.new_sheet.cell(row=194, column=2, value='ONLINE CAMPUS').font = Font(size=9, bold=True)
        for row in (4, 41, 79, 117, 155, 194):
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=4)

    # Checks each row to make sure it has class data
    def row_has_class(self, crn):
        return any(char.isdigit() for char in crn)

    # Adds data to spreadsheet
    def add_data(self, crn, subj, class_name, sec, class_credits, title, days, class_time, cap, act, comments, prof,
                 location, row_type, campus):

        # Online classes are formatted differently from all other classes
        if 'Online' not in campus:
            self.new_sheet.cell(row=row_type, column=2, value=crn)
            self.new_sheet.cell(row=row_type, column=3, value=subj)

            if cap < 19 and cap != 9 and 'MultiCast' not in comments:
                self.new_sheet.cell(row=row_type, column=4, value=f'{class_name}-{sec}*')
            elif 'MultiCast' in comments:
                self.new_sheet.cell(row=row_type, column=4, value=f'{class_name}-{sec}+')
            else:
                self.new_sheet.cell(row=row_type, column=4, value=f'{class_name}-{sec}')

            self.new_sheet.cell(row=row_type, column=5, value=class_credits)
            self.new_sheet.cell(row=row_type, column=6, value=title.upper())
            self.new_sheet.cell(row=row_type, column=7, value=days)
            self.new_sheet.cell(row=row_type, column=8, value=class_time.upper())

            if act != 0:
                self.new_sheet.cell(row=row_type, column=9, value=act)
            self.new_sheet.cell(row=row_type, column=10, value=cap)
            self.new_sheet.cell(row=row_type, column=11, value=location)

            if prof == 'TBA':
                self.new_sheet.cell(row=row_type, column=12, value='STAFF')
            else:
                original_name = prof.split(' ')
                first_initial = original_name[0][0]
                last_name = original_name[-1]
                # Maybe try this with list comprehension???
                if '-' in last_name:
                    # Maybe try combining these two statements???
                    hyphenated_name = last_name.split('-')
                    last_name = hyphenated_name[-1]
                formatted_name = f'{first_initial}. {last_name}'
                self.new_sheet.cell(row=row_type, column=12, value=formatted_name[0:-4].upper())

        # Formatting for Online classes
        else:
            self.new_sheet.cell(row=row_type, column=2, value=crn)
            self.new_sheet.cell(row=row_type, column=3, value=subj)

            if cap < 19 and cap != 9:
                self.new_sheet.cell(row=row_type, column=4, value=f'{class_name}-{sec}*')
            else:
                self.new_sheet.cell(row=row_type, column=4, value=f'{class_name}-{sec}')

            self.new_sheet.cell(row=row_type, column=5, value=class_credits)
            self.new_sheet.cell(row=row_type, column=6, value=title.upper())
            self.new_sheet.cell(row=row_type, column=7, value='')

            if act != 0:
                self.new_sheet.cell(row=row_type, column=9, value=act)
            self.new_sheet.cell(row=row_type, column=10, value=cap)

            if prof == 'TBA':
                self.new_sheet.cell(row=row_type, column=8, value='STAFF')
            else:
                original_name = prof.split(' ')
                first_initial = original_name[0][0]
                last_name = original_name[-1]
                # Maybe try this with list comprehension???
                if '-' in last_name:
                    # Maybe try combining these two statements???
                    hyphenated_name = last_name.split('-')
                    last_name = hyphenated_name[-1]
                formatted_name = f'{first_initial}. {last_name}'
                self.new_sheet.cell(row=row_type, column=8, value=formatted_name[0:-4].upper())

        # The faster less brute-force version if I can get it to work
        # I need it to be able to add the list of values to specific rows, not just to the end of the sheet
        # new_row = [crn, subj, class_name, sec, campus, class_credits, title, days, class_time, cap, act, prof, location]
        # self.new_sheet[self.row].value = new_row
        # self.row += 1
        # print('Row constructed successfully!')

    # Creates data frame and pushes df data to add_data method
    def create_data_frame(self, html_source, file):
        self.create_new_sheet(file)
        # Pulls second to last table from site
        # 'header=0' allows my to properly label each column
        df = pd.read_html(html_source, header=0)[-2]

        for i in range(0, df.shape[0]):
            if self.row_has_class(df.iat[i, self.col_crn]):
                crn = df.iat[i, self.col_crn]
                subj = df.iat[i, self.col_subj]
                class_name = df.iat[i, self.col_class]
                sec = df.iat[i, self.col_sec]
                campus = df.iat[i, self.col_campus]
                class_credits = df.iat[i, self.col_credits]
                title = df.iat[i, self.col_title]
                days = df.iat[i, self.col_days]
                class_time = df.iat[i, self.col_time]
                cap = df.iat[i, self.col_cap]
                act = df.iat[i, self.col_act]
                comments = df.iat[i, self.col_comments]
                prof = df.iat[i, self.col_prof]
                location = df.iat[i, self.col_location]

                if int(cap) > 0:
                    if 'Alpharetta' in campus:
                        self.alp_row += 1
                        row_type = self.alp_row
                        room_number = [i for i in location.split() if i.isdigit()][0]

                        if 'Bldg A' in location:
                            location = f'AA {room_number}'
                        else:
                            location = f'AB {room_number}'
                    elif 'Clarkston' in campus:
                        self.clk_row += 1
                        row_type = self.clk_row
                        room_number = [i for i in location.split() if i.isdigit()][0]

                        if 'Bldg B' in location:
                            location = f'CB {room_number}'
                        elif 'Bldg C' in location:
                            location = f'CC {room_number}'
                        elif 'Bldg D' in location:
                            location = f'CD {room_number}'
                        elif 'Bldg E' in location:
                            location = f'CE {room_number}'
                        else:
                            location = f'CH {room_number}'
                    elif 'Decatur' in campus:
                        self.dec_row += 1
                        row_type = self.dec_row
                        room_number = [i for i in location.split() if i.isdigit()][0]

                        if 'Bldg. SB' in location:
                            location = f'SB {room_number}'
                        else:
                            location = f'SC {room_number}'
                    elif 'Dunwoody' in campus:
                        self.dun_row += 1
                        row_type = self.dun_row
                        room_number = [i for i in location.split() if i.isdigit()][0]

                        if 'Classroom' in location:
                            location = f'E {room_number}'
                        elif 'Science' in location:
                            location = f'C {room_number}'
                        else:
                            location = f'A {room_number}'
                    elif 'Newton' in campus:
                        self.newt_row += 1
                        row_type = self.newt_row
                        room_number = [i for i in location.split() if i.isdigit()][0]
                        location = f'1N {room_number}'
                    else:
                        self.onl_row += 1
                        row_type = self.onl_row
                    print(f'Success: Index: {i}')
                    self.add_data(int(crn), subj, class_name, sec, int(float(class_credits)), title, days, class_time,
                                  int(cap), int(act), comments, prof,
                                  location, row_type, campus)
                else:
                    print(f'Error: Class not found. Index: {i}')
            else:
                print(f'I am not a row. Index: {i}')

        # Sets alignment for every cell after sheet is created all at once
        for row in range(self.new_sheet.min_row, self.new_sheet.max_row + 8):
            self.new_sheet.row_dimensions[row].height = 14.25
            for column in range(self.new_sheet.min_column, self.new_sheet.max_column + 1):
                coordinate = self.new_sheet.cell(row=row, column=column).coordinate
                self.new_sheet[coordinate].alignment = Alignment(horizontal='center')

        # Sets size for specific cells in sheet all at once
        for row in list(chain(range(6, self.alp_row + 1), range(43, self.clk_row + 1), range(81, self.dec_row + 1),
                              range(119, self.dun_row + 1), range(157, self.newt_row + 1), range(196, self.onl_row + 1))):
            coordinate = self.new_sheet.cell(row=row, column=6).coordinate
            self.new_sheet[coordinate].font = Font(name='Arial', size=8)

            for column in list(chain(range(3, 6), range(7, 13))):
                coordinate = self.new_sheet.cell(row=row, column=column).coordinate
                self.new_sheet[coordinate].font = Font(name='Arial', size=9)

        # Sets borders for all cells in tables
        for row in list(chain(range(5, self.alp_row + 1), range(42, self.clk_row + 1), range(80, self.dec_row + 1),
                              range(118, self.dun_row + 1), range(156, self.newt_row + 1), range(195, self.onl_row + 1))):
            if row < 195:
                max_column = self.new_sheet.max_column + 1
            else:
                max_column = self.new_sheet.max_column - 1
            for column in range(self.new_sheet.min_column, max_column):
                coordinate = self.new_sheet.cell(row=row, column=column).coordinate
                self.new_sheet[coordinate].border = self.regular_border

        # List of header rows
        header_rows = [5, 42, 80, 118, 156, 195]
        for row in header_rows:
            for column in range(self.new_sheet.min_column, self.new_sheet.max_column + 1):
                self.new_sheet.row_dimensions[row].height = 21

        footer_tables = [self.alp_row + 2, self.clk_row + 2, self.dec_row + 2, self.dun_row + 2, self. newt_row + 2,
                         self.onl_row + 2]

        for row in footer_tables:
            self.new_sheet.cell(row=row, column=2, value=f'Course ID Legend').border = self.regular_border
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=5)
            self.new_sheet.cell(row=row, column=6, value=f'Host Campus for Multicast Courses').border = self.regular_border
            row += 1

            self.new_sheet.cell(row=row, column=2, value=f'* = Embedded honors class').border = self.legend_border
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=5)
            self.new_sheet.cell(row=row, column=6, value=f'α = Alpharetta').border = self.legend_border
            row += 1

            self.new_sheet.cell(row=row, column=2, value=f'+ = Multicast and embedded honors class').border = \
                self.bottom_legend_border
            self.new_sheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=5)
            self.new_sheet.cell(row=row, column=6, value=f'Σ = Clarkston').border = self.legend_border
            row += 1

            self.new_sheet.cell(row=row, column=6, value=f'∆ = Decatur').border = self.legend_border
            row += 1

            self.new_sheet.cell(row=row, column=6, value=f'λ = Dunwoody').border = self.legend_border
            row += 1

            self.new_sheet.cell(row=row, column=6, value=f'Ω = Newton').border = self.bottom_legend_border
            row += 1

        # Creates list of all numbers that need to be summed
        act_sum = [f'=SUM(I6:I{self.alp_row})', f'=SUM(I43:I{self.clk_row})', f'=SUM(I81:I{self.dec_row})',
                   f'=SUM(I119:I{self.dun_row})', f'=SUM(I157:I{self.newt_row})', f'=SUM(I196:I{self.onl_row})']
        cap_sum = [f'=SUM(J6:J{self.alp_row})', f'=SUM(J43:J{self.clk_row})', f'=SUM(J81:J{self.dec_row})',
                   f'=SUM(J119:J{self.dun_row})', f'=SUM(J157:J{self.newt_row})', f'=SUM(J196:J{self.onl_row})']
        # Used to index through list of numbers that need to be summed
        sum_index = 0

        # Adds to total act and cap to each campus complete w/ borders
        for row in [self.alp_row + 1, self.clk_row + 1, self.dec_row + 1, self.dun_row + 1, self.newt_row + 1,
                    self.onl_row + 1]:
            self.new_sheet.cell(row=row, column=8, value='Total').border = self.regular_border

            self.new_sheet.cell(row=row, column=9, value=act_sum[sum_index]).border = self.regular_border

            self.new_sheet.cell(row=row, column=10, value=cap_sum[sum_index]).border = self.regular_border
            sum_index += 1

        self.book.save(file)

        # *****NOTE - DO NOT DELETE***** This block of code puts data frame into sheet.
        # This does not need to happen every time I run the program while testing, so it is commented out
        # .................................................................................
        # # Appends dataframe to existing sheet via mode='a'
        # with pd.ExcelWriter('Fall Schedule March 25th copy.xlsx', mode='a') as writer:
        #     df.to_excel(writer, sheet_name='DataFrame')
        # .................................................................................
